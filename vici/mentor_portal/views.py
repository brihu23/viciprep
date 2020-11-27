from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.http import HttpResponse, HttpResponseRedirect, Http404
from django.shortcuts import render
from django.core.exceptions import SuspiciousOperation 
from django.urls import reverse
from django import forms
from django.core.paginator import Paginator
import json
from django.http import JsonResponse
from datetime import datetime
from django.shortcuts import redirect
from pyexcel_xls import get_data as xls_get
from pyexcel_xlsx import get_data as xlsx_get
from django.utils.datastructures import MultiValueDictKeyError
import openpyxl
from pathlib import Path
from django import template


def index(request):
    xlsx_file = Path('.', 'file1.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    all_companies = tuple(sheet.rows)
    

    context= {
    	"headers" : all_companies[0],
        "all" : all_companies[1:len(all_companies)],
        "height": "40",
        "range": range(10, 39)
    }      
    return render(request, "mentor_portal/excel_editing.html", context)

def card(request):
    xlsx_file = Path('.', 'file1.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    all_companies = tuple(sheet.rows)
    

    context= {
        "headers" : all_companies[0],
        "all" : all_companies[1:len(all_companies)],
        "height": "40",
        "range": range(10, 39)
    }      
    return render(request, "mentor_portal/card.html", context)

def feature(request):
    xlsx_file = Path('.', 'file12.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    all_companies = tuple(sheet.rows)
    

    context= {
        "headers" : all_companies[0],
        "all" : all_companies[1:len(all_companies)],
        "height": "40",
        "range": range(10, 39)
    }      
    return render(request, "mentor_portal/feature.html", context)